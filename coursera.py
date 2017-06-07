import re
import lxml.etree as etree
import bs4
import requests
from openpyxl import Workbook


def get_courses_list():
    raw_xml = requests.get('https://www.coursera.org/sitemap~www~courses.xml').content
    parser_xml = etree.XMLParser(remove_blank_text=True)
    root_xml = etree.fromstring(raw_xml, parser_xml)
    list_course = []
    for index in range(0,19):
        list_course.append(root_xml[index][0].text)
    return list_course


def get_course_info(url):
    coursera = requests.get(url)
    soup = bs4.BeautifulSoup(coursera.content, "lxml")
    content = str(soup.findAll('script', {'type': 'application/ld+json'}))

    name_course = soup.html.head.title.string
    count_week = len(soup.findAll('div', {'class': 'week-heading body-2-text'}))

    date_start = re.findall('"startDate":"(\d\d\d\d-\d\d-\d\d)', content)
    date_start = date_start[0] if date_start else ''

    lang_course = re.findall('"inLanguage":"(\w\w)', content)
    lang_course = lang_course[0] if lang_course else ''

    rating_course = re.findall('"ratingValue":(\d.\d)', content)
    rating_course = rating_course[0] if rating_course else ''
    return name_course, lang_course, date_start, count_week, rating_course


def output_courses_info_to_xlsx(filepath, list_value):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Name course'
    ws['B1'] = 'language'
    ws['C1'] = 'Date start'
    ws['D1'] = 'Count of week'
    ws['E1'] = 'Rating of course'
    line_sheet = 2
    for course_value in list_value:
        ws['A' + str(line_sheet)] = course_value[0]
        ws['B' + str(line_sheet)] = course_value[1]
        ws['C' + str(line_sheet)] = course_value[2]
        ws['D' + str(line_sheet)] = course_value[3]
        ws['E' + str(line_sheet)] = course_value[4]
        line_sheet += 1
    wb.save(filepath)


if __name__ == '__main__':

    list_value_course = []
    list_url_course = get_courses_list()
    for course in list_url_course:
        list_value_course.append(get_course_info(course))
    output_courses_info_to_xlsx('./sample.xlsx', list_value_course)
